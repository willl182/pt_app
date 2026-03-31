#!/usr/bin/env python3
"""
Add an Algorithm A worksheet to validation_calculations.xlsx.

The sheet follows ISO 13528:2022 Annex C.3 using iterative winsorization,
not the deprecated Huber-weighting approach.
"""
import csv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = "/home/w182/w421/pt_app/validation"
DATA_DIR = "/home/w182/w421/pt_app/data"

header_fill = PatternFill("solid", fgColor="4472C4")
section_fill = PatternFill("solid", fgColor="D9EAD3")
input_fill = PatternFill("solid", fgColor="FFF2CC")
result_fill = PatternFill("solid", fgColor="E2EFDA")
winsor_fill = PatternFill("solid", fgColor="D9EAF7")
header_font = Font(bold=True, color="FFFFFF")
formula_font = Font(color="0000FF")
bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

MAX_ITER = 10
TOL = 1e-6


def style_cell(cell, fill=None, font=None, center=False):
    """Apply common formatting to a cell."""
    cell.border = thin_border
    if fill is not None:
      cell.fill = fill
    if font is not None:
      cell.font = font
    if center:
      cell.alignment = Alignment(horizontal="center", vertical="center")


def load_summary_data(pollutant, level):
    """Load aggregated participant means for a pollutant/level pair."""
    rows = []
    with open(f"{DATA_DIR}/summary_n4.csv", "r", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            if row["participant_id"] == "ref":
                continue
            if row["pollutant"] != pollutant or row["level"] != level:
                continue
            try:
                rows.append((row["participant_id"], float(row["mean_value"])))
            except ValueError:
                continue

    rows.sort(key=lambda item: item[0])
    return rows


def add_algorithm_a_sheet(wb):
    """Create a winsorization-based Algorithm A validation sheet."""
    if "Algorithm_A" in wb.sheetnames:
        del wb["Algorithm_A"]
    ws = wb.create_sheet("Algorithm_A")

    participants = load_summary_data("so2", "60-nmol/mol")
    n = len(participants)
    if n < 3:
        raise ValueError("Algorithm_A example requires at least 3 observations")

    values_start = 12
    values_end = values_start + n - 1

    ws["A1"] = "ALGORITHM A - ROBUST MEAN AND STANDARD DEVIATION"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:M1")

    ws["A2"] = "ISO 13528:2022 Annex C.3 - Iterative winsorization"
    ws.merge_cells("A2:M2")

    ws["A3"] = "Example: SO2 60-nmol/mol participant results"
    ws.merge_cells("A3:M3")

    ws["A5"] = "PARAMETERS"
    ws["A5"].font = bold_font

    parameter_rows = [
        ("Convergence tolerance", TOL, None),
        ("Max iterations", MAX_ITER, None),
        ("MAD scale factor", 1.483, "Used in s*_0 = 1.483 × MAD"),
        ("Winsorization factor", 1.5, "delta = 1.5 × s*"),
        ("Scale adjustment factor", 1.134, "Applied to winsorized sample SD"),
    ]
    for offset, (label, value, note) in enumerate(parameter_rows, start=6):
        ws.cell(row=offset, column=1, value=label)
        ws.cell(row=offset, column=2, value=value)
        style_cell(ws.cell(row=offset, column=2), fill=input_fill)
        if note:
            ws.cell(row=offset, column=3, value=note)

    ws["A11"] = "INPUT DATA"
    ws["A11"].font = bold_font

    headers = ["i", "participant_id", "xi", "|xi - median|"]
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=values_start - 1, column=col, value=label)
        style_cell(cell, fill=header_fill, font=header_font, center=True)

    for idx, (participant_id, value) in enumerate(participants, start=values_start):
        ws.cell(row=idx, column=1, value=idx - values_start + 1)
        ws.cell(row=idx, column=2, value=participant_id)
        ws.cell(row=idx, column=3, value=value)
        style_cell(ws.cell(row=idx, column=1), center=True)
        style_cell(ws.cell(row=idx, column=2))
        style_cell(ws.cell(row=idx, column=3))

    init_row = values_end + 2
    ws.cell(row=init_row, column=1, value="INITIAL ESTIMATES (Iteration 0)")
    ws.cell(row=init_row, column=1).font = bold_font

    median_cell = f"B{init_row + 1}"
    mad_cell = f"B{init_row + 2}"
    s0_cell = f"B{init_row + 3}"

    ws.cell(row=init_row + 1, column=1, value="x*_0 = median(xi)")
    ws.cell(
        row=init_row + 1,
        column=2,
        value=f"=MEDIAN(C{values_start}:C{values_end})"
    )
    style_cell(ws.cell(row=init_row + 1, column=2), fill=result_fill, font=formula_font)

    for row in range(values_start, values_end + 1):
        ws.cell(row=row, column=4, value=f"=ABS(C{row}-{median_cell})")
        style_cell(ws.cell(row=row, column=4), font=formula_font)

    ws.cell(row=init_row + 2, column=1, value="MAD = median(|xi - x*_0|)")
    ws.cell(
        row=init_row + 2,
        column=2,
        value=f"=MEDIAN(D{values_start}:D{values_end})"
    )
    style_cell(ws.cell(row=init_row + 2, column=2), font=formula_font)

    ws.cell(row=init_row + 3, column=1, value="s*_0 = 1.483 × MAD")
    ws.cell(row=init_row + 3, column=2, value=f"=$B$8*{mad_cell}")
    style_cell(ws.cell(row=init_row + 3, column=2), fill=result_fill, font=formula_font)

    iter_header_row = init_row + 6
    ws.cell(row=iter_header_row - 1, column=1, value="ITERATION SUMMARY")
    ws.cell(row=iter_header_row - 1, column=1).font = bold_font

    summary_headers = [
        "Iter", "x*_prev", "s*_prev", "delta", "lower", "upper",
        "n_winsorized", "x*_new", "s*_new", "delta_x", "delta_s",
        "delta_max", "Converged?"
    ]
    for col, label in enumerate(summary_headers, start=1):
        cell = ws.cell(row=iter_header_row, column=col, value=label)
        style_cell(cell, fill=header_fill, font=header_font, center=True)

    iter_start = iter_header_row + 1
    iter_end = iter_start + MAX_ITER - 1

    for row in range(iter_start, iter_end + 1):
        iter_no = row - iter_start + 1
        ws.cell(row=row, column=1, value=iter_no)
        style_cell(ws.cell(row=row, column=1), center=True)

        if iter_no == 1:
            ws.cell(row=row, column=2, value=f"={median_cell}")
            ws.cell(row=row, column=3, value=f"={s0_cell}")
        else:
            ws.cell(row=row, column=2, value=f"=H{row - 1}")
            ws.cell(row=row, column=3, value=f"=I{row - 1}")

        ws.cell(row=row, column=4, value=f"=$B$9*C{row}")
        ws.cell(row=row, column=5, value=f"=B{row}-D{row}")
        ws.cell(row=row, column=6, value=f"=B{row}+D{row}")
        ws.cell(
            row=row,
            column=7,
            value=(
                f'=COUNTIF(N{row}:R{row},"YES")'
                if n <= 5 else
                f'=COUNTIF(N{row}:{chr(ord("N") + n - 1)}{row},"YES")'
            )
        )
        for col in range(2, 8):
            style_cell(ws.cell(row=row, column=col), font=formula_font)

    detail_header_row = iter_end + 3
    ws.cell(row=detail_header_row - 1, column=1, value="DETAIL BY ITERATION")
    ws.cell(row=detail_header_row - 1, column=1).font = bold_font

    detail_headers = ["Iter"]
    for idx, (participant_id, _) in enumerate(participants, start=1):
        detail_headers.extend([
            f"{participant_id}_xi*",
            f"{participant_id}_winsor?"
        ])

    for col, label in enumerate(detail_headers, start=1):
        cell = ws.cell(row=detail_header_row, column=col, value=label)
        style_cell(cell, fill=header_fill, font=header_font, center=True)

    detail_start = detail_header_row + 1
    for iter_row in range(iter_start, iter_end + 1):
        out_row = detail_start + (iter_row - iter_start)
        ws.cell(row=out_row, column=1, value=iter_row - iter_start + 1)
        style_cell(ws.cell(row=out_row, column=1), center=True)

        detail_col = 2
        for value_row in range(values_start, values_end + 1):
            xi_cell = f"$C${value_row}"
            lower_cell = f"$E{iter_row}"
            upper_cell = f"$F{iter_row}"
            wins_cell = ws.cell(
                row=out_row,
                column=detail_col,
                value=f"=MAX(MIN({xi_cell},{upper_cell}),{lower_cell})"
            )
            style_cell(wins_cell, fill=winsor_fill, font=formula_font)

            flag_cell = ws.cell(
                row=out_row,
                column=detail_col + 1,
                value=f'=IF(OR({xi_cell}<{lower_cell},{xi_cell}>{upper_cell}),"YES","NO")'
            )
            style_cell(flag_cell, font=formula_font, center=True)

            detail_col += 2

        first_wins_col = 2
        last_wins_col = 2 + 2 * n - 2
        wins_cells = ",".join(
            f"{column_letter(col)}{out_row}"
            for col in range(first_wins_col, last_wins_col + 1, 2)
        )
        ws.cell(row=iter_row, column=8, value=f"=AVERAGE({wins_cells})")
        ws.cell(row=iter_row, column=9, value=f"=$B$10*STDEV.S({wins_cells})")
        ws.cell(row=iter_row, column=10, value=f"=ABS(H{iter_row}-B{iter_row})")
        ws.cell(row=iter_row, column=11, value=f"=ABS(I{iter_row}-C{iter_row})")
        ws.cell(row=iter_row, column=12, value=f"=MAX(J{iter_row},K{iter_row})")
        ws.cell(row=iter_row, column=13, value=f'=IF(L{iter_row}<$B$6,"YES","NO")')
        for col in range(8, 14):
            style_cell(ws.cell(row=iter_row, column=col), fill=result_fill, font=formula_font)

    final_row = detail_start + MAX_ITER + 2
    ws.cell(row=final_row, column=1, value="FINAL RESULTS")
    ws.cell(row=final_row, column=1).font = bold_font

    ws.cell(row=final_row + 1, column=1, value="x* (last iteration)")
    ws.cell(row=final_row + 1, column=2, value=f"=H{iter_end}")
    style_cell(ws.cell(row=final_row + 1, column=2), fill=result_fill, font=formula_font)

    ws.cell(row=final_row + 2, column=1, value="s* (last iteration)")
    ws.cell(row=final_row + 2, column=2, value=f"=I{iter_end}")
    style_cell(ws.cell(row=final_row + 2, column=2), fill=result_fill, font=formula_font)

    ws.cell(row=final_row + 3, column=1, value="u(x_pt) = 1.25 × s* / sqrt(n)")
    ws.cell(row=final_row + 3, column=2, value=f"=1.25*B{final_row + 2}/SQRT(COUNTA(C{values_start}:C{values_end}))")
    style_cell(ws.cell(row=final_row + 3, column=2), fill=result_fill, font=formula_font)

    ref_row = final_row + 6
    ws.cell(row=ref_row, column=1, value="FORMULA REFERENCE")
    ws.cell(row=ref_row, column=1).font = bold_font

    formulas = [
        ("Step 0", "x*_0 = median(xi)"),
        ("Step 0", "s*_0 = 1.483 × median(|xi - x*_0|)"),
        ("Step 1", "delta = 1.5 × s*"),
        ("Step 2", "xi* = clamp(xi, x* - delta, x* + delta)"),
        ("Step 3", "x*_new = mean(xi*)"),
        ("Step 4", "s*_new = 1.134 × STDEV.S(xi*)"),
        ("Stop", "Converges when max(|Δx*|, |Δs*|) < 1e-6"),
    ]
    for offset, (step, formula) in enumerate(formulas, start=1):
        ws.cell(row=ref_row + offset, column=1, value=step)
        ws.cell(row=ref_row + offset, column=2, value=formula)

    widths = {
        "A": 28, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 15,
        "H": 18, "I": 18, "J": 15, "K": 15, "L": 15, "M": 15
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for col in range(14, 14 + 2 * n):
        ws.column_dimensions[column_letter(col)].width = 16


def column_letter(index):
    """Convert a 1-based column index to Excel letter(s)."""
    letters = []
    while index > 0:
        index, rem = divmod(index - 1, 26)
        letters.append(chr(65 + rem))
    return "".join(reversed(letters))


def main():
    xlsx_path = f"{SCRIPT_DIR}/validation_calculations.xlsx"
    print(f"Loading {xlsx_path}...")
    wb = load_workbook(xlsx_path)

    print("Adding Algorithm_A sheet...")
    add_algorithm_a_sheet(wb)

    wb.save(xlsx_path)
    print(f"Saved: {xlsx_path}")
    print("Sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
