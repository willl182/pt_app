#!/usr/bin/env python3
"""
Add Algorithm A iteration sheet to validation_calculations.xlsx
ISO 13528:2022 Annex C - Robust Mean and Standard Deviation
"""
import csv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = "/home/w182/w421/pt_app/validation"
DATA_DIR = "/home/w182/w421/pt_app/data"

header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, color="FFFFFF")
formula_font = Font(color="0000FF")
input_fill = PatternFill("solid", fgColor="FFF2CC")
result_fill = PatternFill("solid", fgColor="E2EFDA")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def load_summary_data(pollutant, level):
    """Load mean_value for a specific pollutant/level from summary_n4.csv"""
    values = []
    with open(f"{DATA_DIR}/summary_n4.csv", 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row['pollutant'] == pollutant and row['level'] == level:
                try:
                    values.append(float(row['mean_value']))
                except ValueError:
                    pass
    return values

def add_algorithm_a_sheet(wb):
    """Add Algorithm A iteration sheet demonstrating the robust estimation process"""
    if "Algorithm_A" in wb.sheetnames:
        del wb["Algorithm_A"]
    ws = wb.create_sheet("Algorithm_A")
    
    # Load SO2 60-nmol/mol data (same as other sheets)
    values = load_summary_data("so2", "60-nmol/mol")
    n = len(values)
    
    # Title
    ws['A1'] = "ALGORITHM A - ROBUST MEAN AND STANDARD DEVIATION"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:K1')
    
    ws['A2'] = "ISO 13528:2022 Annex C - Iterative estimation with Huber weights"
    ws.merge_cells('A2:K2')
    
    ws['A3'] = "Example: SO2 60-nmol/mol participant results"
    ws.merge_cells('A3:K3')
    
    # Parameters section
    ws['A5'] = "PARAMETERS"
    ws['A5'].font = Font(bold=True)
    
    ws['A6'] = "Convergence tolerance"
    ws['B6'] = 0.001
    ws['B6'].fill = input_fill
    
    ws['A7'] = "Max iterations"
    ws['B7'] = 50
    ws['B7'].fill = input_fill
    
    ws['A8'] = "Huber tuning constant"
    ws['B8'] = 1.5
    ws['B8'].fill = input_fill
    ws['C8'] = "(used in u = (x - x*) / (1.5 × s*))"
    
    # Data section - Column A: values, Column B: |xi - median|
    ws['A10'] = "INPUT DATA"
    ws['A10'].font = Font(bold=True)
    
    headers = ["i", "xi", "|xi - median|"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    data_start = 12
    for i, v in enumerate(values):
        ws.cell(row=data_start + i, column=1, value=i + 1)
        ws.cell(row=data_start + i, column=2, value=v)
    data_end = data_start + n - 1
    
    # Initial calculations section
    init_row = data_end + 2
    ws.cell(row=init_row, column=1, value="INITIAL ESTIMATES (Iteration 0)")
    ws.cell(row=init_row, column=1).font = Font(bold=True)
    
    ws.cell(row=init_row + 1, column=1, value="n (count)")
    ws.cell(row=init_row + 1, column=2, value=f"=COUNT(B{data_start}:B{data_end})")
    ws.cell(row=init_row + 1, column=2).font = formula_font
    
    ws.cell(row=init_row + 2, column=1, value="Median (x*₀)")
    ws.cell(row=init_row + 2, column=2, value=f"=MEDIAN(B{data_start}:B{data_end})")
    ws.cell(row=init_row + 2, column=2).font = formula_font
    ws.cell(row=init_row + 2, column=2).fill = result_fill
    median_cell = f"B{init_row + 2}"
    
    # Fill in |xi - median| column
    for i in range(n):
        cell = ws.cell(row=data_start + i, column=3, value=f"=ABS(B{data_start + i}-{median_cell})")
        cell.font = formula_font
    
    ws.cell(row=init_row + 3, column=1, value="MAD")
    ws.cell(row=init_row + 3, column=2, value=f"=MEDIAN(C{data_start}:C{data_end})")
    ws.cell(row=init_row + 3, column=2).font = formula_font
    
    ws.cell(row=init_row + 4, column=1, value="MADe (s*₀ = 1.483 × MAD)")
    ws.cell(row=init_row + 4, column=2, value=f"=1.483*B{init_row + 3}")
    ws.cell(row=init_row + 4, column=2).font = formula_font
    ws.cell(row=init_row + 4, column=2).fill = result_fill
    s0_cell = f"B{init_row + 4}"
    
    # Iteration table section
    iter_row = init_row + 6
    ws.cell(row=iter_row, column=1, value="ITERATION TABLE")
    ws.cell(row=iter_row, column=1).font = Font(bold=True)
    
    # Headers for iteration calculations
    iter_headers = ["i", "xi", "u = (xi-x*)/1.5s*", "w (weight)", "w×xi", "w×(xi-x*)²"]
    for col, h in enumerate(iter_headers, 1):
        cell = ws.cell(row=iter_row + 1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Iteration 1 calculations - use initial estimates
    iter1_start = iter_row + 2
    x_star_ref = median_cell  # First iteration uses median
    s_star_ref = s0_cell  # First iteration uses MADe
    
    for i in range(n):
        r = iter1_start + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=f"=B{data_start + i}")
        ws.cell(row=r, column=2).font = formula_font
        
        # u = (xi - x*) / (1.5 × s*)
        ws.cell(row=r, column=3, value=f"=(B{r}-{x_star_ref})/(1.5*{s_star_ref})")
        ws.cell(row=r, column=3).font = formula_font
        
        # w = IF(|u|<=1, 1, 1/u²)
        ws.cell(row=r, column=4, value=f"=IF(ABS(C{r})<=1,1,1/(C{r}^2))")
        ws.cell(row=r, column=4).font = formula_font
        
        # w × xi
        ws.cell(row=r, column=5, value=f"=D{r}*B{r}")
        ws.cell(row=r, column=5).font = formula_font
        
        # w × (xi - x*)² - will reference the updated x* after sums
        ws.cell(row=r, column=6, value=f"=D{r}*(B{r}-{x_star_ref})^2")
        ws.cell(row=r, column=6).font = formula_font
    
    iter1_end = iter1_start + n - 1
    
    # Sums and new estimates
    sum_row = iter1_end + 1
    ws.cell(row=sum_row, column=1, value="Σ")
    ws.cell(row=sum_row, column=1).font = Font(bold=True)
    ws.cell(row=sum_row, column=4, value=f"=SUM(D{iter1_start}:D{iter1_end})")
    ws.cell(row=sum_row, column=4).font = formula_font
    ws.cell(row=sum_row, column=5, value=f"=SUM(E{iter1_start}:E{iter1_end})")
    ws.cell(row=sum_row, column=5).font = formula_font
    ws.cell(row=sum_row, column=6, value=f"=SUM(F{iter1_start}:F{iter1_end})")
    ws.cell(row=sum_row, column=6).font = formula_font
    
    # New estimates after iteration 1
    result_row = sum_row + 2
    ws.cell(row=result_row, column=1, value="ITERATION 1 RESULTS")
    ws.cell(row=result_row, column=1).font = Font(bold=True)
    
    ws.cell(row=result_row + 1, column=1, value="x*₁ = Σ(w×x) / Σw")
    ws.cell(row=result_row + 1, column=2, value=f"=E{sum_row}/D{sum_row}")
    ws.cell(row=result_row + 1, column=2).font = formula_font
    ws.cell(row=result_row + 1, column=2).fill = result_fill
    x1_cell = f"B{result_row + 1}"
    
    ws.cell(row=result_row + 2, column=1, value="s*₁ = √(Σ(w×(x-x*)²) / Σw)")
    ws.cell(row=result_row + 2, column=2, value=f"=SQRT(F{sum_row}/D{sum_row})")
    ws.cell(row=result_row + 2, column=2).font = formula_font
    ws.cell(row=result_row + 2, column=2).fill = result_fill
    s1_cell = f"B{result_row + 2}"
    
    ws.cell(row=result_row + 3, column=1, value="Δx* = |x*₁ - x*₀|")
    ws.cell(row=result_row + 3, column=2, value=f"=ABS({x1_cell}-{x_star_ref})")
    ws.cell(row=result_row + 3, column=2).font = formula_font
    
    ws.cell(row=result_row + 4, column=1, value="Δs* = |s*₁ - s*₀|")
    ws.cell(row=result_row + 4, column=2, value=f"=ABS({s1_cell}-{s_star_ref})")
    ws.cell(row=result_row + 4, column=2).font = formula_font
    
    ws.cell(row=result_row + 5, column=1, value="Converged? (max(Δ) < tol)")
    ws.cell(row=result_row + 5, column=2, value=f'=IF(MAX(B{result_row+3},B{result_row+4})<$B$6,"YES","NO")')
    ws.cell(row=result_row + 5, column=2).font = formula_font
    ws.cell(row=result_row + 5, column=2).fill = result_fill
    
    # Formula reference section
    ref_row = result_row + 8
    ws.cell(row=ref_row, column=1, value="ALGORITHM A FORMULA REFERENCE")
    ws.cell(row=ref_row, column=1).font = Font(bold=True)
    
    formulas = [
        ("Step 1 - Initialize:", "x* = median(x), s* = 1.483 × MAD"),
        ("Step 2 - Compute u:", "uᵢ = (xᵢ - x*) / (1.5 × s*)"),
        ("Step 3 - Compute weights:", "wᵢ = 1 if |uᵢ| ≤ 1, else wᵢ = 1/uᵢ²"),
        ("Step 4 - Update x*:", "x* = Σ(wᵢ × xᵢ) / Σwᵢ"),
        ("Step 5 - Update s*:", "s* = √(Σ(wᵢ × (xᵢ - x*)²) / Σwᵢ)"),
        ("Step 6 - Convergence:", "Stop when |Δx*| < tol AND |Δs*| < tol"),
    ]
    
    for i, (name, formula) in enumerate(formulas):
        ws.cell(row=ref_row + 1 + i, column=1, value=name)
        ws.cell(row=ref_row + 1 + i, column=2, value=formula)
    
    # Interpretation section
    interp_row = ref_row + len(formulas) + 3
    ws.cell(row=interp_row, column=1, value="INTERPRETATION")
    ws.cell(row=interp_row, column=1).font = Font(bold=True)
    
    notes = [
        "• Huber weights down-weight outliers (w < 1 for |u| > 1)",
        "• Final x* is the robust assigned value (x_pt)",
        "• Final s* can be used to derive σ_pt",
        "• Typically converges in 5-15 iterations",
        "• Reference: ISO 13528:2022, Annex C",
    ]
    for i, note in enumerate(notes):
        ws.cell(row=interp_row + 1 + i, column=1, value=note)
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 18

def main():
    xlsx_path = f"{SCRIPT_DIR}/validation_calculations.xlsx"
    print(f"Loading {xlsx_path}...")
    wb = load_workbook(xlsx_path)
    
    print("Adding Algorithm_A sheet...")
    add_algorithm_a_sheet(wb)
    
    wb.save(xlsx_path)
    print(f"Saved: {xlsx_path}")
    print("Sheets:", wb.sheetnames)

if __name__ == "__main__":
    main()
