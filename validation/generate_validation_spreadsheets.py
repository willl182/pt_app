#!/usr/bin/env python3
"""
Validation Spreadsheet Generator for PT App Calculations
=========================================================
Generates Excel spreadsheets with formulas to validate all calculations
in app.R and the ptcalc/ package per ISO 13528:2022.

Formulas validated:
- Homogeneity: ss, sw, s_x_bar_sq, criterion c = 0.3 * sigma_pt
- Stability: diff_hom_stab, criterion comparison
- Robust Statistics: MADe = 1.483 * MAD, nIQR = 0.7413 * IQR
- PT Scores: z, z', zeta, En scores

Uses only standard library + openpyxl (no pandas dependency).
"""

import csv
import os
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    print("Alternatively, run this script in an R environment with:")
    print('  reticulate::py_install("openpyxl")')
    exit(1)

# Configuration
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, '..', 'data')
OUTPUT_DIR = SCRIPT_DIR

# Styles
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='4472C4')
FORMULA_FONT = Font(color='0000FF')  # Blue for formulas
INPUT_FILL = PatternFill('solid', fgColor='FFF2CC')  # Light yellow for inputs
RESULT_FILL = PatternFill('solid', fgColor='E2EFDA')  # Light green for results
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def load_csv(filename):
    """Load CSV file and return list of dicts."""
    filepath = os.path.join(DATA_DIR, filename)
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        return list(reader)


def filter_data(data, pollutant, level):
    """Filter data by pollutant and level."""
    return [row for row in data if row['pollutant'] == pollutant and row['level'] == level]


def pivot_homogeneity_data(data):
    """Pivot homogeneity data to get replicates as columns per sample_id."""
    samples = defaultdict(dict)
    for row in data:
        sample_id = int(row['sample_id'])
        replicate = int(row['replicate'])
        value = float(row['value'])
        samples[sample_id][f'rep{replicate}'] = value
    
    result = []
    for sample_id in sorted(samples.keys()):
        result.append({
            'sample_id': sample_id,
            'rep1': samples[sample_id].get('rep1', 0),
            'rep2': samples[sample_id].get('rep2', 0)
        })
    return result


def style_header_row(ws, row, max_col):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws):
    """Auto-adjust column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def create_homogeneity_sheet(wb, hom_data):
    """
    Create homogeneity validation sheet.
    
    Formulas (ISO 13528:2022 Section 9.2):
    - sample_mean = AVERAGE(rep1, rep2)
    - grand_mean = AVERAGE(sample_means)
    - s_x_bar_sq = VAR.S(sample_means)
    - range = ABS(rep1 - rep2)
    - sw = SQRT(SUMSQ(ranges) / (2 * g))
    - ss_sq = ABS(s_x_bar_sq - sw^2/m)
    - ss = SQRT(ss_sq)
    - c = 0.3 * sigma_pt
    """
    ws = wb.create_sheet("Homogeneity")
    
    # Filter for SO2 60-nmol/mol
    example = filter_data(hom_data, 'so2', '60-nmol/mol')
    pivot = pivot_homogeneity_data(example)
    g = len(pivot)  # number of samples
    m = 2  # number of replicates
    
    # Title
    ws['A1'] = "HOMOGENEITY VALIDATION - SO2 60-nmol/mol"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    ws['A2'] = "ISO 13528:2022 Section 9.2 - Between-sample and within-sample standard deviations"
    ws.merge_cells('A2:H2')
    
    # Parameters section
    ws['A4'] = "Parameters"
    ws['A4'].font = Font(bold=True)
    ws['A5'] = "g (samples)"
    ws['B5'] = g
    ws['A6'] = "m (replicates)"
    ws['B6'] = m
    ws['A7'] = "σ_pt (user input)"
    ws['B7'] = 0.6  # Example sigma_pt for SO2 60 nmol/mol
    ws['B7'].fill = INPUT_FILL
    
    # Data table header
    headers = ['Sample ID', 'Replicate 1', 'Replicate 2', 'Sample Mean', 'Range |R1-R2|', 'Range²']
    start_row = 9
    for col, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=header)
    style_header_row(ws, start_row, len(headers))
    
    # Data rows with formulas
    data_start = start_row + 1
    for i, row in enumerate(pivot, start=data_start):
        r = i
        ws.cell(row=r, column=1, value=row['sample_id'])
        ws.cell(row=r, column=2, value=row['rep1'])
        ws.cell(row=r, column=3, value=row['rep2'])
        # Sample Mean formula
        ws.cell(row=r, column=4, value=f'=AVERAGE(B{r},C{r})')
        ws.cell(row=r, column=4).font = FORMULA_FONT
        # Range formula
        ws.cell(row=r, column=5, value=f'=ABS(B{r}-C{r})')
        ws.cell(row=r, column=5).font = FORMULA_FONT
        # Range squared formula
        ws.cell(row=r, column=6, value=f'=E{r}^2')
        ws.cell(row=r, column=6).font = FORMULA_FONT
    
    data_end = data_start + g - 1
    
    # Summary statistics section
    summary_row = data_end + 2
    ws.cell(row=summary_row, column=1, value="CALCULATED STATISTICS")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    
    # Grand Mean
    ws.cell(row=summary_row+1, column=1, value="Grand Mean (x̄̄)")
    ws.cell(row=summary_row+1, column=2, value=f'=AVERAGE(D{data_start}:D{data_end})')
    ws.cell(row=summary_row+1, column=2).font = FORMULA_FONT
    ws.cell(row=summary_row+1, column=2).fill = RESULT_FILL
    
    # Variance of sample means
    ws.cell(row=summary_row+2, column=1, value="s²_x̄ (Var of means)")
    ws.cell(row=summary_row+2, column=2, value=f'=VAR.S(D{data_start}:D{data_end})')
    ws.cell(row=summary_row+2, column=2).font = FORMULA_FONT
    ws.cell(row=summary_row+2, column=2).fill = RESULT_FILL
    
    # Sum of squared ranges
    ws.cell(row=summary_row+3, column=1, value="Σ(Range²)")
    ws.cell(row=summary_row+3, column=2, value=f'=SUM(F{data_start}:F{data_end})')
    ws.cell(row=summary_row+3, column=2).font = FORMULA_FONT
    
    # sw (within-sample SD)
    ws.cell(row=summary_row+4, column=1, value="sw (within-sample SD)")
    ws.cell(row=summary_row+4, column=2, value=f'=SQRT(B{summary_row+3}/(2*B5))')
    ws.cell(row=summary_row+4, column=2).font = FORMULA_FONT
    ws.cell(row=summary_row+4, column=2).fill = RESULT_FILL
    ws.cell(row=summary_row+4, column=3, value="Formula: √(Σw²/(2g))")
    
    # sw²
    ws.cell(row=summary_row+5, column=1, value="sw²")
    ws.cell(row=summary_row+5, column=2, value=f'=B{summary_row+4}^2')
    ws.cell(row=summary_row+5, column=2).font = FORMULA_FONT
    
    # ss² (between-sample variance)
    ws.cell(row=summary_row+6, column=1, value="ss² (|s²_x̄ - sw²/m|)")
    ws.cell(row=summary_row+6, column=2, value=f'=ABS(B{summary_row+2}-B{summary_row+5}/B6)')
    ws.cell(row=summary_row+6, column=2).font = FORMULA_FONT
    ws.cell(row=summary_row+6, column=2).fill = RESULT_FILL
    ws.cell(row=summary_row+6, column=3, value="Formula: |s²_x̄ - sw²/m|")
    
    # ss (between-sample SD)
    ws.cell(row=summary_row+7, column=1, value="ss (between-sample SD)")
    ws.cell(row=summary_row+7, column=2, value=f'=SQRT(B{summary_row+6})')
    ws.cell(row=summary_row+7, column=2).font = FORMULA_FONT
    ws.cell(row=summary_row+7, column=2).fill = RESULT_FILL
    ws.cell(row=summary_row+7, column=3, value="Formula: √ss²")
    
    # Criterion section
    crit_row = summary_row + 9
    ws.cell(row=crit_row, column=1, value="HOMOGENEITY CRITERION")
    ws.cell(row=crit_row, column=1).font = Font(bold=True)
    
    ws.cell(row=crit_row+1, column=1, value="c = 0.3 × σ_pt")
    ws.cell(row=crit_row+1, column=2, value='=0.3*B7')
    ws.cell(row=crit_row+1, column=2).font = FORMULA_FONT
    ws.cell(row=crit_row+1, column=2).fill = RESULT_FILL
    
    ws.cell(row=crit_row+2, column=1, value="ss ≤ c ?")
    ws.cell(row=crit_row+2, column=2, value=f'=IF(B{summary_row+7}<=B{crit_row+1},"PASS","FAIL")')
    ws.cell(row=crit_row+2, column=2).font = FORMULA_FONT
    ws.cell(row=crit_row+2, column=2).fill = RESULT_FILL
    
    auto_width(ws)
    return ws


def create_stability_sheet(wb, hom_data, stab_data):
    """
    Create stability validation sheet.
    
    Formulas (ISO 13528:2022 Section 9.3):
    - diff_hom_stab = |stability_mean - homogeneity_mean|
    - criterion: diff ≤ 0.3 × σ_pt
    """
    ws = wb.create_sheet("Stability")
    
    # Filter for SO2 60-nmol/mol
    hom_ex = filter_data(hom_data, 'so2', '60-nmol/mol')
    stab_ex = filter_data(stab_data, 'so2', '60-nmol/mol')
    
    hom_pivot = pivot_homogeneity_data(hom_ex)
    stab_pivot = pivot_homogeneity_data(stab_ex)
    
    # Title
    ws['A1'] = "STABILITY VALIDATION - SO2 60-nmol/mol"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "ISO 13528:2022 Section 9.3 - Stability assessment"
    ws.merge_cells('A2:F2')
    
    # Parameters
    ws['A4'] = "Parameters"
    ws['A4'].font = Font(bold=True)
    ws['A5'] = "σ_pt (user input)"
    ws['B5'] = 0.6
    ws['B5'].fill = INPUT_FILL
    
    # Homogeneity data
    ws['A7'] = "HOMOGENEITY DATA"
    ws['A7'].font = Font(bold=True)
    
    headers = ['Sample ID', 'Rep 1', 'Rep 2', 'Mean']
    for col, h in enumerate(headers, 1):
        ws.cell(row=8, column=col, value=h)
    style_header_row(ws, 8, len(headers))
    
    row = 9
    for i, r in enumerate(hom_pivot):
        ws.cell(row=row+i, column=1, value=r['sample_id'])
        ws.cell(row=row+i, column=2, value=r['rep1'])
        ws.cell(row=row+i, column=3, value=r['rep2'])
        ws.cell(row=row+i, column=4, value=f'=AVERAGE(B{row+i},C{row+i})')
        ws.cell(row=row+i, column=4).font = FORMULA_FONT
    
    hom_end = row + len(hom_pivot) - 1
    
    ws.cell(row=hom_end+1, column=1, value="Homogeneity Grand Mean")
    ws.cell(row=hom_end+1, column=1).font = Font(bold=True)
    ws.cell(row=hom_end+1, column=4, value=f'=AVERAGE(D{row}:D{hom_end})')
    ws.cell(row=hom_end+1, column=4).font = FORMULA_FONT
    ws.cell(row=hom_end+1, column=4).fill = RESULT_FILL
    hom_mean_cell = f'D{hom_end+1}'
    
    # Stability data
    stab_start = hom_end + 4
    ws.cell(row=stab_start, column=1, value="STABILITY DATA")
    ws.cell(row=stab_start, column=1).font = Font(bold=True)
    
    for col, h in enumerate(headers, 1):
        ws.cell(row=stab_start+1, column=col, value=h)
    style_header_row(ws, stab_start+1, len(headers))
    
    row = stab_start + 2
    for i, r in enumerate(stab_pivot):
        ws.cell(row=row+i, column=1, value=r['sample_id'])
        ws.cell(row=row+i, column=2, value=r['rep1'])
        ws.cell(row=row+i, column=3, value=r['rep2'])
        ws.cell(row=row+i, column=4, value=f'=AVERAGE(B{row+i},C{row+i})')
        ws.cell(row=row+i, column=4).font = FORMULA_FONT
    
    stab_end = row + len(stab_pivot) - 1
    
    ws.cell(row=stab_end+1, column=1, value="Stability Grand Mean")
    ws.cell(row=stab_end+1, column=1).font = Font(bold=True)
    ws.cell(row=stab_end+1, column=4, value=f'=AVERAGE(D{row}:D{stab_end})')
    ws.cell(row=stab_end+1, column=4).font = FORMULA_FONT
    ws.cell(row=stab_end+1, column=4).fill = RESULT_FILL
    stab_mean_cell = f'D{stab_end+1}'
    
    # Stability assessment
    assess_row = stab_end + 4
    ws.cell(row=assess_row, column=1, value="STABILITY ASSESSMENT")
    ws.cell(row=assess_row, column=1).font = Font(bold=True)
    
    ws.cell(row=assess_row+1, column=1, value="|Stab Mean - Hom Mean|")
    ws.cell(row=assess_row+1, column=2, value=f'=ABS({stab_mean_cell}-{hom_mean_cell})')
    ws.cell(row=assess_row+1, column=2).font = FORMULA_FONT
    ws.cell(row=assess_row+1, column=2).fill = RESULT_FILL
    
    ws.cell(row=assess_row+2, column=1, value="Criterion c = 0.3 × σ_pt")
    ws.cell(row=assess_row+2, column=2, value='=0.3*B5')
    ws.cell(row=assess_row+2, column=2).font = FORMULA_FONT
    ws.cell(row=assess_row+2, column=2).fill = RESULT_FILL
    
    ws.cell(row=assess_row+3, column=1, value="diff ≤ c ?")
    ws.cell(row=assess_row+3, column=2, value=f'=IF(B{assess_row+1}<=B{assess_row+2},"PASS","FAIL")')
    ws.cell(row=assess_row+3, column=2).font = FORMULA_FONT
    ws.cell(row=assess_row+3, column=2).fill = RESULT_FILL
    
    # Additional stability uncertainty calculations (new in app.R)
    ws.cell(row=assess_row+5, column=1, value="ADDITIONAL STABILITY CALCULATIONS")
    ws.cell(row=assess_row+5, column=1).font = Font(bold=True)
    
    # u_stab = Dmax / sqrt(3)
    ws.cell(row=assess_row+6, column=1, value="u_stab = Dmax / √3")
    ws.cell(row=assess_row+6, column=2, value=f'=B{assess_row+1}/SQRT(3)')
    ws.cell(row=assess_row+6, column=2).font = FORMULA_FONT
    ws.cell(row=assess_row+6, column=2).fill = RESULT_FILL
    ws.cell(row=assess_row+6, column=3, value="← Uncertainty due to instability")
    
    # Expanded criterion (from app.R compute_stability_metrics)
    ws.cell(row=assess_row+8, column=1, value="EXPANDED CRITERION")
    ws.cell(row=assess_row+8, column=1).font = Font(bold=True)
    
    # Calculate SD of all hom and stab values
    all_hom_values = [r['rep1'] for r in hom_pivot] + [r['rep2'] for r in hom_pivot]
    all_stab_values = [r['rep1'] for r in stab_pivot] + [r['rep2'] for r in stab_pivot]
    
    import statistics
    sd_hom = statistics.stdev(all_hom_values) if len(all_hom_values) > 1 else 0
    sd_stab = statistics.stdev(all_stab_values) if len(all_stab_values) > 1 else 0
    
    ws.cell(row=assess_row+9, column=1, value="SD(hom data)")
    ws.cell(row=assess_row+9, column=2, value=sd_hom)
    
    ws.cell(row=assess_row+10, column=1, value="n_hom")
    ws.cell(row=assess_row+10, column=2, value=len(all_hom_values))
    
    ws.cell(row=assess_row+11, column=1, value="u_hom_mean = SD/√n_hom")
    ws.cell(row=assess_row+11, column=2, value=f'=B{assess_row+9}/SQRT(B{assess_row+10})')
    ws.cell(row=assess_row+11, column=2).font = FORMULA_FONT
    
    ws.cell(row=assess_row+12, column=1, value="SD(stab data)")
    ws.cell(row=assess_row+12, column=2, value=sd_stab)
    
    ws.cell(row=assess_row+13, column=1, value="n_stab")
    ws.cell(row=assess_row+13, column=2, value=len(all_stab_values))
    
    ws.cell(row=assess_row+14, column=1, value="u_stab_mean = SD/√n_stab")
    ws.cell(row=assess_row+14, column=2, value=f'=B{assess_row+12}/SQRT(B{assess_row+13})')
    ws.cell(row=assess_row+14, column=2).font = FORMULA_FONT
    
    ws.cell(row=assess_row+15, column=1, value="c_expanded = c + 2×√(u_hom_mean² + u_stab_mean²)")
    ws.cell(row=assess_row+15, column=2, value=f'=B{assess_row+2}+2*SQRT(B{assess_row+11}^2+B{assess_row+14}^2)')
    ws.cell(row=assess_row+15, column=2).font = FORMULA_FONT
    ws.cell(row=assess_row+15, column=2).fill = RESULT_FILL
    
    ws.cell(row=assess_row+16, column=1, value="diff ≤ c_expanded ?")
    ws.cell(row=assess_row+16, column=2, value=f'=IF(B{assess_row+1}<=B{assess_row+15},"PASS","FAIL")')
    ws.cell(row=assess_row+16, column=2).font = FORMULA_FONT
    ws.cell(row=assess_row+16, column=2).fill = RESULT_FILL
    
    auto_width(ws)
    return ws


def create_robust_stats_sheet(wb, summary_data):
    """
    Create robust statistics validation sheet.
    
    Formulas (ISO 13528:2022 Section 9.4):
    - nIQR = 0.7413 × (Q3 - Q1)
    - MADe = 1.483 × MEDIAN(|xi - MEDIAN(x)|)
    """
    ws = wb.create_sheet("Robust_Stats")
    
    # Filter example data: SO2 60-nmol/mol
    example = filter_data(summary_data, 'so2', '60-nmol/mol')
    values = [float(row['mean_value']) for row in example if row['mean_value']]
    n = len(values)
    
    # Title
    ws['A1'] = "ROBUST STATISTICS VALIDATION - SO2 60-nmol/mol"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "ISO 13528:2022 Section 9.4 - MADe and nIQR calculations"
    ws.merge_cells('A2:F2')
    
    # Data column headers
    ws['A4'] = "Values (xi)"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "|xi - median|"
    ws['B4'].font = Font(bold=True)
    
    for i, v in enumerate(values, start=5):
        ws.cell(row=i, column=1, value=v)
    
    data_end = 5 + n - 1
    
    # Calculations section
    calc_row = data_end + 3
    ws.cell(row=calc_row, column=1, value="CALCULATIONS")
    ws.cell(row=calc_row, column=1).font = Font(bold=True)
    
    # n
    ws.cell(row=calc_row+1, column=1, value="n (count)")
    ws.cell(row=calc_row+1, column=2, value=f'=COUNT(A5:A{data_end})')
    ws.cell(row=calc_row+1, column=2).font = FORMULA_FONT
    
    # Median
    ws.cell(row=calc_row+2, column=1, value="Median")
    ws.cell(row=calc_row+2, column=2, value=f'=MEDIAN(A5:A{data_end})')
    ws.cell(row=calc_row+2, column=2).font = FORMULA_FONT
    ws.cell(row=calc_row+2, column=2).fill = RESULT_FILL
    median_cell = f'B{calc_row+2}'
    
    # Q1
    ws.cell(row=calc_row+3, column=1, value="Q1 (25th percentile)")
    ws.cell(row=calc_row+3, column=2, value=f'=QUARTILE.INC(A5:A{data_end},1)')
    ws.cell(row=calc_row+3, column=2).font = FORMULA_FONT
    
    # Q3
    ws.cell(row=calc_row+4, column=1, value="Q3 (75th percentile)")
    ws.cell(row=calc_row+4, column=2, value=f'=QUARTILE.INC(A5:A{data_end},3)')
    ws.cell(row=calc_row+4, column=2).font = FORMULA_FONT
    
    # IQR
    ws.cell(row=calc_row+5, column=1, value="IQR (Q3 - Q1)")
    ws.cell(row=calc_row+5, column=2, value=f'=B{calc_row+4}-B{calc_row+3}')
    ws.cell(row=calc_row+5, column=2).font = FORMULA_FONT
    
    # nIQR
    ws.cell(row=calc_row+6, column=1, value="nIQR = 0.7413 × IQR")
    ws.cell(row=calc_row+6, column=2, value=f'=0.7413*B{calc_row+5}')
    ws.cell(row=calc_row+6, column=2).font = FORMULA_FONT
    ws.cell(row=calc_row+6, column=2).fill = RESULT_FILL
    ws.cell(row=calc_row+6, column=3, value="← Robust SD estimator")
    
    # Add absolute deviations column
    for i, v in enumerate(values, start=5):
        ws.cell(row=i, column=2, value=f'=ABS(A{i}-{median_cell})')
        ws.cell(row=i, column=2).font = FORMULA_FONT
    
    # MAD (median of absolute deviations)
    ws.cell(row=calc_row+7, column=1, value="MAD (median of |xi-median|)")
    ws.cell(row=calc_row+7, column=2, value=f'=MEDIAN(B5:B{data_end})')
    ws.cell(row=calc_row+7, column=2).font = FORMULA_FONT
    
    # MADe
    ws.cell(row=calc_row+8, column=1, value="MADe = 1.483 × MAD")
    ws.cell(row=calc_row+8, column=2, value=f'=1.483*B{calc_row+7}')
    ws.cell(row=calc_row+8, column=2).font = FORMULA_FONT
    ws.cell(row=calc_row+8, column=2).fill = RESULT_FILL
    ws.cell(row=calc_row+8, column=3, value="← Robust SD estimator")
    
    # Standard mean and SD for comparison
    ws.cell(row=calc_row+10, column=1, value="COMPARISON (Classical)")
    ws.cell(row=calc_row+10, column=1).font = Font(bold=True)
    
    ws.cell(row=calc_row+11, column=1, value="Mean")
    ws.cell(row=calc_row+11, column=2, value=f'=AVERAGE(A5:A{data_end})')
    ws.cell(row=calc_row+11, column=2).font = FORMULA_FONT
    
    ws.cell(row=calc_row+12, column=1, value="SD")
    ws.cell(row=calc_row+12, column=2, value=f'=STDEV.S(A5:A{data_end})')
    ws.cell(row=calc_row+12, column=2).font = FORMULA_FONT
    
    auto_width(ws)
    return ws


def create_algorithm_a_sheet(wb, summary_data):
    """
    Create Algorithm A validation sheet.
    
    Algorithm A (ISO 13528:2022 Annex C):
    1. Initialize: x* = median, s* = 1.483 × MAD
    2. u = (xi - x*) / (1.5 × s*)
    3. weights: w = 1 if |u| ≤ 1, else w = 1/u²
    4. x*_new = Σ(w×xi) / Σw
    5. s*_new = √(Σ(w×(xi-x*)²) / Σw)
    6. Iterate until convergence
    """
    ws = wb.create_sheet("Algorithm_A")
    
    # Filter example data: SO2 60-nmol/mol
    example = filter_data(summary_data, 'so2', '60-nmol/mol')
    values = [float(row['mean_value']) for row in example if row['mean_value']][:12]
    n = len(values)
    
    # Title
    ws['A1'] = "ALGORITHM A VALIDATION - SO2 60-nmol/mol"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    ws['A2'] = "ISO 13528:2022 Annex C - Iterative robust estimation"
    ws.merge_cells('A2:H2')
    
    ws['A3'] = "Note: Shows first iteration. Full algorithm requires multiple iterations until convergence."
    ws['A3'].font = Font(italic=True)
    ws.merge_cells('A3:H3')
    
    # Initial estimates
    ws['A5'] = "INITIAL ESTIMATES"
    ws['A5'].font = Font(bold=True)
    
    # Data with headers
    headers = ['i', 'xi', '|xi - median|', 'u = (xi-x*)/(1.5×s*)', 'weight w', 'w×xi', 'w×(xi-x*)²']
    for col, h in enumerate(headers, 1):
        ws.cell(row=7, column=col, value=h)
    style_header_row(ws, 7, len(headers))
    
    # Data rows
    for i, v in enumerate(values, start=8):
        ws.cell(row=i, column=1, value=i-7)  # index
        ws.cell(row=i, column=2, value=v)    # xi
    
    data_end = 8 + n - 1
    
    # Initial x* and s* (median and MADe)
    init_row = data_end + 2
    ws.cell(row=init_row, column=1, value="x* (initial) = Median")
    ws.cell(row=init_row, column=2, value=f'=MEDIAN(B8:B{data_end})')
    ws.cell(row=init_row, column=2).font = FORMULA_FONT
    ws.cell(row=init_row, column=2).fill = RESULT_FILL
    x_star_cell = f'B{init_row}'
    
    ws.cell(row=init_row+1, column=1, value="s* (initial) = 1.483×MAD")
    # First calculate MAD
    for i in range(8, data_end+1):
        ws.cell(row=i, column=3, value=f'=ABS(B{i}-{x_star_cell})')
        ws.cell(row=i, column=3).font = FORMULA_FONT
    
    ws.cell(row=init_row+1, column=2, value=f'=1.483*MEDIAN(C8:C{data_end})')
    ws.cell(row=init_row+1, column=2).font = FORMULA_FONT
    ws.cell(row=init_row+1, column=2).fill = RESULT_FILL
    s_star_cell = f'B{init_row+1}'
    
    # Calculate u values
    for i in range(8, data_end+1):
        ws.cell(row=i, column=4, value=f'=(B{i}-{x_star_cell})/(1.5*{s_star_cell})')
        ws.cell(row=i, column=4).font = FORMULA_FONT
    
    # Calculate weights: w = IF(ABS(u)<=1, 1, 1/u^2)
    for i in range(8, data_end+1):
        ws.cell(row=i, column=5, value=f'=IF(ABS(D{i})<=1,1,1/(D{i}^2))')
        ws.cell(row=i, column=5).font = FORMULA_FONT
    
    # w × xi
    for i in range(8, data_end+1):
        ws.cell(row=i, column=6, value=f'=E{i}*B{i}')
        ws.cell(row=i, column=6).font = FORMULA_FONT
    
    # w × (xi - x*)²
    for i in range(8, data_end+1):
        ws.cell(row=i, column=7, value=f'=E{i}*(B{i}-{x_star_cell})^2')
        ws.cell(row=i, column=7).font = FORMULA_FONT
    
    # Iteration 1 results
    iter_row = init_row + 4
    ws.cell(row=iter_row, column=1, value="ITERATION 1 RESULTS")
    ws.cell(row=iter_row, column=1).font = Font(bold=True)
    
    ws.cell(row=iter_row+1, column=1, value="Σw")
    ws.cell(row=iter_row+1, column=2, value=f'=SUM(E8:E{data_end})')
    ws.cell(row=iter_row+1, column=2).font = FORMULA_FONT
    sum_w_cell = f'B{iter_row+1}'
    
    ws.cell(row=iter_row+2, column=1, value="Σ(w×xi)")
    ws.cell(row=iter_row+2, column=2, value=f'=SUM(F8:F{data_end})')
    ws.cell(row=iter_row+2, column=2).font = FORMULA_FONT
    
    ws.cell(row=iter_row+3, column=1, value="x*_new = Σ(w×xi)/Σw")
    ws.cell(row=iter_row+3, column=2, value=f'=B{iter_row+2}/{sum_w_cell}')
    ws.cell(row=iter_row+3, column=2).font = FORMULA_FONT
    ws.cell(row=iter_row+3, column=2).fill = RESULT_FILL
    ws.cell(row=iter_row+3, column=3, value="← Robust Mean (after iter 1)")
    
    ws.cell(row=iter_row+4, column=1, value="Σ(w×(xi-x*)²)")
    ws.cell(row=iter_row+4, column=2, value=f'=SUM(G8:G{data_end})')
    ws.cell(row=iter_row+4, column=2).font = FORMULA_FONT
    
    ws.cell(row=iter_row+5, column=1, value="s*_new = √(Σ(w×(xi-x*)²)/Σw)")
    ws.cell(row=iter_row+5, column=2, value=f'=SQRT(B{iter_row+4}/{sum_w_cell})')
    ws.cell(row=iter_row+5, column=2).font = FORMULA_FONT
    ws.cell(row=iter_row+5, column=2).fill = RESULT_FILL
    ws.cell(row=iter_row+5, column=3, value="← Robust SD (after iter 1)")
    
    auto_width(ws)
    return ws


def create_pt_scores_sheet(wb):
    """
    Create PT scores validation sheet.
    
    Formulas (ISO 13528:2022 Section 10):
    - z = (x - x_pt) / σ_pt
    - z' = (x - x_pt) / √(σ_pt² + u_xpt_def²)
    - ζ = (x - x_pt) / √(u_x² + u_xpt_def²)
    - En = (x - x_pt) / √(U_x² + U_xpt²)
    - u_xpt_def = √(u_xpt² + u_hom² + u_stab²)
    """
    ws = wb.create_sheet("PT_Scores")
    
    # Title
    ws['A1'] = "PT SCORES VALIDATION"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    ws['A2'] = "ISO 13528:2022 Section 10 - Performance scores"
    ws.merge_cells('A2:H2')
    
    # Parameters section
    ws['A4'] = "ASSIGNED VALUE & PARAMETERS (Example: SO2 60-nmol/mol)"
    ws['A4'].font = Font(bold=True)
    
    params = [
        ('x_pt (assigned value)', 59.9),
        ('σ_pt (std dev for PT)', 0.6),
        ('u_xpt (std uncertainty of x_pt)', 0.1),
        ('u_hom (uncertainty from homogeneity)', 0.05),
        ('u_stab (uncertainty from stability)', 0.03),
        ('u_xpt_def (definitive uncertainty)', None),
        ('U_xpt (expanded uncertainty, k=2)', None),
    ]
    
    for i, (name, val) in enumerate(params, start=5):
        ws.cell(row=i, column=1, value=name)
        if val is not None:
            ws.cell(row=i, column=2, value=val)
            ws.cell(row=i, column=2).fill = INPUT_FILL
    
    # u_xpt_def = sqrt(u_xpt^2 + u_hom^2 + u_stab^2)
    ws.cell(row=10, column=2, value='=SQRT(B7^2+B8^2+B9^2)')
    ws.cell(row=10, column=2).font = FORMULA_FONT
    ws.cell(row=10, column=2).fill = RESULT_FILL
    ws.cell(row=10, column=3, value="← √(u_xpt² + u_hom² + u_stab²)")
    
    # U_xpt = k * u_xpt_def (k=2)
    ws.cell(row=11, column=2, value='=2*B10')
    ws.cell(row=11, column=2).font = FORMULA_FONT
    ws.cell(row=11, column=2).fill = RESULT_FILL
    
    # Example participants
    ws['A13'] = "PARTICIPANT RESULTS"
    ws['A13'].font = Font(bold=True)
    
    headers = ['Participant', 'x (result)', 'u_x (std unc)', 'U_x (exp unc)', 
               'z-score', 'z\'-score', 'ζ-score', 'En-score', 'z Eval', 'En Eval']
    for col, h in enumerate(headers, 1):
        ws.cell(row=14, column=col, value=h)
    style_header_row(ws, 14, len(headers))
    
    # Sample participant data (realistic values for SO2 ~60 nmol/mol)
    participants = [
        ('Lab A', 59.95, 0.15, 0.30),
        ('Lab B', 59.80, 0.20, 0.40),
        ('Lab C', 60.50, 0.10, 0.20),
        ('Lab D', 59.90, 0.25, 0.50),
        ('Lab E', 58.50, 0.30, 0.60),
    ]
    
    for i, (name, x, u_x, U_x) in enumerate(participants, start=15):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=x)
        ws.cell(row=i, column=2).fill = INPUT_FILL
        ws.cell(row=i, column=3, value=u_x)
        ws.cell(row=i, column=3).fill = INPUT_FILL
        ws.cell(row=i, column=4, value=U_x)
        ws.cell(row=i, column=4).fill = INPUT_FILL
        
        # z-score: (x - x_pt) / σ_pt
        ws.cell(row=i, column=5, value=f'=(B{i}-$B$5)/$B$6')
        ws.cell(row=i, column=5).font = FORMULA_FONT
        
        # z'-score: (x - x_pt) / √(σ_pt² + u_xpt_def²) - uses B10 for u_xpt_def
        ws.cell(row=i, column=6, value=f'=(B{i}-$B$5)/SQRT($B$6^2+$B$10^2)')
        ws.cell(row=i, column=6).font = FORMULA_FONT
        
        # ζ-score: (x - x_pt) / √(u_x² + u_xpt_def²) - uses B10 for u_xpt_def
        ws.cell(row=i, column=7, value=f'=(B{i}-$B$5)/SQRT(C{i}^2+$B$10^2)')
        ws.cell(row=i, column=7).font = FORMULA_FONT
        
        # En-score: (x - x_pt) / √(U_x² + U_xpt²) - uses B11 for U_xpt
        ws.cell(row=i, column=8, value=f'=(B{i}-$B$5)/SQRT(D{i}^2+$B$11^2)')
        ws.cell(row=i, column=8).font = FORMULA_FONT
        
        # z evaluation
        ws.cell(row=i, column=9, value=f'=IF(ABS(E{i})<=2,"Satisfactorio",IF(ABS(E{i})<3,"Cuestionable","No satisfactorio"))')
        ws.cell(row=i, column=9).font = FORMULA_FONT
        
        # En evaluation
        ws.cell(row=i, column=10, value=f'=IF(ABS(H{i})<=1,"Satisfactorio","No satisfactorio")')
        ws.cell(row=i, column=10).font = FORMULA_FONT
    
    # Formula reference section
    ref_row = 15 + len(participants) + 2
    ws.cell(row=ref_row, column=1, value="FORMULA REFERENCE")
    ws.cell(row=ref_row, column=1).font = Font(bold=True)
    
    formulas = [
        ('z-score', 'z = (x - x_pt) / σ_pt'),
        ('z\'-score', 'z\' = (x - x_pt) / √(σ_pt² + u_xpt_def²)'),
        ('ζ-score (zeta)', 'ζ = (x - x_pt) / √(u_x² + u_xpt_def²)'),
        ('En-score', 'En = (x - x_pt) / √(U_x² + U_xpt²)'),
        ('u_xpt_def', 'u_xpt_def = √(u_xpt² + u_hom² + u_stab²)'),
    ]
    
    for i, (name, formula) in enumerate(formulas, start=ref_row+1):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=formula)
    
    # Evaluation criteria
    eval_row = ref_row + len(formulas) + 2
    ws.cell(row=eval_row, column=1, value="EVALUATION CRITERIA")
    ws.cell(row=eval_row, column=1).font = Font(bold=True)
    
    criteria = [
        ('z-score', '|z| ≤ 2: Satisfactorio, 2 < |z| < 3: Cuestionable, |z| ≥ 3: No satisfactorio'),
        ('En-score', '|En| ≤ 1: Satisfactorio, |En| > 1: No satisfactorio'),
    ]
    
    for i, (name, crit) in enumerate(criteria, start=eval_row+1):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=crit)
    
    auto_width(ws)
    return ws


def main():
    """Generate all validation spreadsheets."""
    print("Loading data files...")
    
    # Load data
    hom_data = load_csv('homogeneity.csv')
    stab_data = load_csv('stability.csv')
    summary_data = load_csv('summary_n4.csv')
    
    print(f"Loaded homogeneity: {len(hom_data)} rows")
    print(f"Loaded stability: {len(stab_data)} rows")
    print(f"Loaded summary_n4: {len(summary_data)} rows")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    default = wb.active
    if default is not None:
        wb.remove(default)
    
    # Create all sheets
    print("Creating Homogeneity sheet...")
    create_homogeneity_sheet(wb, hom_data)
    
    print("Creating Stability sheet...")
    create_stability_sheet(wb, hom_data, stab_data)
    
    print("Creating Robust Stats sheet...")
    create_robust_stats_sheet(wb, summary_data)
    
    print("Creating Algorithm A sheet...")
    create_algorithm_a_sheet(wb, summary_data)
    
    print("Creating PT Scores sheet...")
    create_pt_scores_sheet(wb)
    
    # Save workbook
    output_path = os.path.join(OUTPUT_DIR, 'validation_calculations.xlsx')
    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    
    print("\n" + "="*60)
    print("NEXT STEP: Run recalc.py to calculate formula values")
    print("="*60)
    print(f"python /home/w182/.config/opencode/skill/xlsx/recalc.py {output_path}")
    
    return output_path


if __name__ == '__main__':
    main()
